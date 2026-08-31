"""GX2-A: Build canonical field matrix from Excel masters + live codebase reality."""
from __future__ import annotations

import csv
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(r"g:\momentra_v2")
FIELD_XLSX = Path(
    r"c:\Users\HI\Downloads\Momentra_Group_217_Screen_4210_Widget_FINAL_Field_Level_Master_Mapping.xlsx"
)
TABLE_XLSX = Path(r"c:\Users\HI\Downloads\Momentra_Master_Table_Repository.xlsx")
OUT_DIR = ROOT / "docs" / "implementation"
OUT_CSV = OUT_DIR / "GX2_FIELD_MATRIX.csv"
OUT_MD = OUT_DIR / "GX2_FIELD_MATRIX.md"
OUT_GAP = OUT_DIR / "GX2_A_GAP_BACKLOG.md"

# Live /v1 contracts known from router.ts (Aug 2026) — overrides stale Excel CONTRACT GAP notes.
LIVE_READ = {
    "GET /v1/me",
    "GET /v1/group/moments",
    "GET /v1/group/moments/:momentId/pulse",
    "GET /v1/group/moments/:momentId/life",
    "GET /v1/group/moments/:momentId/memory",
    "GET /v1/group/moments/:momentId/finance",
    "GET /v1/group/moments/:momentId/activity",
    "GET /v1/group/moments/:momentId/participants",
    "GET /v1/group/invites/:code",
}
LIVE_WRITE = {
    "POST /v1/moments",
    "PATCH /v1/group/moments/:momentId/budget",
    "POST /v1/moments/:momentId/group-expenses",
    "POST /v1/moments/:momentId/contributions",
    "POST /v1/moments/:momentId/settlements",
    "POST /v1/group/invites",
    "POST /v1/group/invites/:code/redeem",
}

# Explicit GX-2 deferred / figmagap / local patterns
DEFERRED_PATTERNS = [
    r"\bupi\b",
    r"google pay",
    r"gpay",
    r"bank transfer",
    r"\bchat\b",
    r"\bfx\b",
    r"exchange rate",
    r"multi-?currency",
    r"remove participant",
    r"shared goal",
    r"community coordination",
]
FIGMA_GAP_PATTERNS = [
    r"\bai\b",
    r"health score",
    r"group pulse score",
    r"intelligence",
    r"coming soon",
]
LOCAL_ONLY_PATTERNS = [
    r"payment rhythm",
    r"join approval",
    r"expense reminders",
    r"photo reminders",
    r"notify",
    r"reminder",
    r"cadence",
    r"split style",
]
SCHEMA_GAP_HINTS = [
    r"no explicit canonical",
    r"no explicit field",
    r"not present in supplied schema",
    r"n/a",
]

# Tables that exist but lack live mounted routes (API_GAP writers)
API_GAP_TABLES = {
    "collaboration.planning_item",
    "collaboration.booking",
    "collaboration.group_update",
    "collaboration.purchase_item",
    "collaboration.resident",
    "collaboration.living_rule",
    "collaboration.shared_asset",
    "collaboration.maintenance_record",
    "collaboration.group_vendor",
    "collaboration.attendance",
    "collaboration.poll",  # transitional; shared.poll is target
    "shared.poll",
    "memory.memory",
    "work.task",
    "work.goal",
}


def norm(s: object) -> str:
    if s is None:
        return ""
    return str(s).strip()


def normalize_api(api: str) -> str:
    a = norm(api)
    if not a or a.upper() in {"N/A", "LOCAL_ONLY", "TBD", "NONE"}:
        return ""
    # unify path params
    a = re.sub(r":[a-zA-Z_]+Id", lambda m: m.group(0), a)
    a = a.replace("{id}", ":momentId").replace("{momentId}", ":momentId")
    # map old personal-style expense to group-expenses where notes say group
    return a


def matches_any(text: str, patterns: list[str]) -> bool:
    t = text.lower()
    return any(re.search(p, t, re.I) for p in patterns)


def family_from_batch(batch: str, gtype: str) -> str:
    b = batch.lower()
    if "purchase" in b or "purchase" in gtype.lower():
        return "Shared Purchase"
    if "living" in b or "living" in gtype.lower() or "flatmate" in gtype.lower():
        return "Shared Living"
    if "common" in b:
        return "Common Group"
    return "Shared Experience"


def subtype_from(gtype: str) -> str:
    g = gtype.replace("—", "-").replace("–", "-")
    if "Trip" in g:
        return "TRIP"
    if "Wedding" in g:
        return "WEDDING"
    if "House Party" in g or "Party" in g:
        return "HOUSE_PARTY"
    if "Office" in g:
        return "OFFICE_OUTING"
    if "Gift" in g:
        return "GIFT_POOL"
    if "Group Purchase" in g:
        return "GROUP_PURCHASE"
    if "Shared Asset" in g or "Asset" in g:
        return "SHARED_ASSET"
    if "Flatmate" in g:
        return "FLATMATES"
    if "Family" in g:
        return "FAMILY_HOUSEHOLD"
    if "Co-living" in g or "Co living" in g or "Coliving" in g:
        return "CO_LIVING"
    if "Custom" in g and "Purchase" in g:
        return "COMMUNITY_PURCHASE"
    if "Custom" in g and "Living" in g:
        return "COMMUNITY_LIVING"
    if "Generic" in g:
        return "GENERIC"
    if "Common" in g:
        return "COMMON"
    return g.split("-")[-1].strip() if "-" in g else g[:40]


def classify_row(row: dict) -> str:
    """Map Excel + live reality → exactly one GX-2 status. No UNKNOWN."""
    excel_status = norm(row["Mapping Status"]).upper()
    api = normalize_api(row["API / Service Contract"])
    api_status = norm(row["API Contract Status"]).upper()
    entity = norm(row["Canonical Entity / Projection"]).lower()
    fields = norm(row["Primary Field(s)"])
    notes = norm(row["Gap / Developer Notes"])
    widget = norm(row["Widget Name"])
    label = norm(row["UI Label / Display"])
    direction = norm(row["Data Direction"]).upper()
    data_beh = norm(row["Data Behaviour"]).upper()
    blob = " ".join([widget, label, notes, api, entity, fields])

    # Explicit DEFERRED product boundaries
    if matches_any(blob, DEFERRED_PATTERNS):
        # multi-currency / payment rails etc.
        if matches_any(blob, [r"multi-?currency", r"\bfx\b", r"exchange rate"]):
            return "DEFERRED"
        if matches_any(blob, [r"\bupi\b", r"google pay", r"gpay", r"bank transfer", r"\bchat\b"]):
            return "DEFERRED"
        if matches_any(blob, [r"remove participant", r"shared goal", r"community coordination"]):
            return "DEFERRED"

    if matches_any(blob, FIGMA_GAP_PATTERNS) and excel_status in {"GAP", "PARTIAL", "STATIC"}:
        # AI insights / invented health — FIGMA_GAP unless already STATIC chrome
        if matches_any(blob, [r"\bai\b", r"health score", r"intelligence"]):
            return "FIGMA_GAP"

    # LOCAL_ONLY / STATIC chrome
    if excel_status == "STATIC" or direction == "STATIC" or api.upper() == "LOCAL_ONLY":
        if matches_any(blob, LOCAL_ONLY_PATTERNS):
            return "LOCAL_ONLY"
        return "LOCAL_ONLY"

    if matches_any(blob, LOCAL_ONLY_PATTERNS) and excel_status in {"GAP", "PARTIAL"}:
        return "LOCAL_ONLY"

    # SCHEMA_GAP — no table
    if excel_status == "GAP" and (
        entity in {"", "n/a"}
        or fields.upper() in {"", "N/A"}
        or matches_any(notes, SCHEMA_GAP_HINTS)
    ):
        return "SCHEMA_GAP"

    # Live API overrides for Excel PARTIAL/CONTRACT GAP
    live_hit = False
    for live in LIVE_READ | LIVE_WRITE:
        # fuzzy: excel may say "GET /v1/group/moments/:momentId/pulse" or partial
        key = live.split(" ", 1)[-1]
        if key and key.lower() in api.lower():
            live_hit = True
            break
    # Special cases Excel didn't know about
    if "budget" in blob.lower() and ("finance.budget" in entity or "budget" in api.lower()):
        live_hit = True  # PATCH budget live
    if "settlement" in blob.lower() and "finance.settlement" in entity:
        live_hit = True
    if "invite" in blob.lower() or "participant" in blob.lower():
        if "moment_participant" in entity or "moment_invite" in entity or "invite" in api.lower():
            live_hit = True
    if "expense" in blob.lower() and ("finance.expense" in entity or "group-expense" in api.lower() or "expenses" in api.lower()):
        live_hit = True
    if "contribution" in blob.lower() and "contribution" in entity:
        live_hit = True

    # API_GAP — table exists, route missing
    table_hit = any(t in entity for t in API_GAP_TABLES)
    if table_hit and not live_hit:
        # planning/booking/poll/update/purchase/resident/memory
        return "API_GAP"

    if excel_status == "GAP" and table_hit:
        return "API_GAP"

    if excel_status == "MAPPED" and live_hit:
        # May still be CLIENT_FIX if notes say client incomplete — default WIRED for mapped+live
        if matches_any(notes, [r"client", r"apk", r"ios", r"force equal", r"coming soon", r"disabled"]):
            return "CLIENT_FIX"
        return "WIRED"

    if excel_status == "PARTIAL":
        if live_hit:
            return "CLIENT_FIX"  # backend ready; binding/nav incomplete
        if table_hit or entity not in {"", "n/a"}:
            return "API_GAP"
        return "SCHEMA_GAP"

    if excel_status == "MAPPED":
        # Mapped in Excel but route may be projection-only or stale
        if api_status in {"CONTRACT GAP", "GAP"} and not live_hit:
            if entity not in {"", "n/a"}:
                return "API_GAP"
            return "SCHEMA_GAP"
        if live_hit or api_status in {"FROZEN HTTP", "FROZEN HTTP / PHYSICAL TRANSITION"}:
            return "WIRED"
        if entity not in {"", "n/a"}:
            return "API_GAP"
        return "LOCAL_ONLY"

    if excel_status == "GAP":
        if table_hit or entity not in {"", "n/a"}:
            return "API_GAP"
        return "SCHEMA_GAP"

    # Fallback — still no UNKNOWN
    if direction in {"READ", "WRITE", "READ/WRITE"} and entity not in {"", "n/a"}:
        return "API_GAP" if not live_hit else "CLIENT_FIX"
    return "LOCAL_ONLY"


def load_field_rows() -> list[dict]:
    wb = load_workbook(FIELD_XLSX, read_only=True, data_only=True)
    ws = wb["01 FINAL Group Field Mapping"]
    rows_iter = ws.iter_rows(values_only=True)
    headers = [norm(h) for h in next(rows_iter)]
    out = []
    for raw in rows_iter:
        if not raw or raw[0] is None:
            continue
        d = {headers[i]: raw[i] if i < len(raw) else None for i in range(len(headers))}
        out.append(d)
    wb.close()
    return out


def load_table_index() -> set[str]:
    """Optional: collect fully-qualified tables from Master Table Repository if present."""
    tables: set[str] = set()
    if not TABLE_XLSX.exists():
        return tables
    wb = load_workbook(TABLE_XLSX, read_only=True, data_only=True)
    # Prefer canonical objects sheet inside field workbook already; try first sheet with Table column
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True, max_row=5))
        if not rows:
            continue
        headers = [norm(h).lower() for h in rows[0]]
        if "fully qualified table" in headers or ("schema" in headers and "table" in headers):
            # read all
            all_rows = list(ws.iter_rows(values_only=True))
            hdr = [norm(h) for h in all_rows[0]]
            for r in all_rows[1:]:
                d = {hdr[i]: r[i] if i < len(r) else None for i in range(len(hdr))}
                fq = norm(d.get("Fully Qualified Table") or "")
                if not fq:
                    sch = norm(d.get("Schema") or "")
                    tbl = norm(d.get("Table") or "")
                    if sch and tbl:
                        fq = f"{sch}.{tbl}"
                if fq and "." in fq:
                    tables.add(fq.lower())
            break
    # Also from field workbook canonical objects
    wb2 = load_workbook(FIELD_XLSX, read_only=True, data_only=True)
    if "13 G-M0 03 Canonical Objects" in wb2.sheetnames:
        ws = wb2["13 G-M0 03 Canonical Objects"]
        rows = list(ws.iter_rows(values_only=True))
        hdr = [norm(h) for h in rows[0]]
        for r in rows[1:]:
            d = {hdr[i]: r[i] if i < len(r) else None for i in range(len(hdr))}
            fq = norm(d.get("Fully Qualified Table") or "")
            if fq:
                tables.add(fq.lower())
    wb2.close()
    wb.close()
    return tables


def android_binding_hint(row: dict, status: str) -> str:
    screen = norm(row["Functional Screen Name"]).lower()
    widget = norm(row["Widget Name"]).lower()
    if "setup" in screen:
        return "GroupExperienceSetupContent / GroupSectionSetupContent"
    if "pulse" in screen:
        return "GroupPulseActiveContent / WeddingPulseActiveContent"
    if "moments" in screen and "list" in screen:
        return "GET group/moments inventory (needs list UI)"
    if "moments" in screen:
        return "GroupMomentsActiveContent"
    if "memory" in screen:
        return "GroupMemoryActiveContent"
    if "action center" in screen or "quick" in screen:
        return "GroupQuickAddHub / GroupActionRegistry"
    if "finance" in screen or "split" in screen:
        return "GroupFinanceScreens"
    if "expense" in screen:
        return "GroupExpenseSheet"
    if "settlement" in screen or "settle" in widget:
        return "GroupSettlementSheet (CTA disabled → CLIENT_FIX)"
    if "contribution" in screen:
        return "GroupContributionSheet"
    if "budget" in screen or "budget" in widget:
        return "GroupBudgetSheet"
    if "activity" in screen:
        return "GroupPulseActiveContent activity section"
    if "plan" in screen or "booking" in screen or "poll" in screen:
        return "Quick Add gap tile (apiGap=true)"
    if status == "LOCAL_ONLY":
        return "Local UI / shell chrome"
    return "See AppShellScreen Group routing"


def ios_binding_hint(row: dict, status: str) -> str:
    a = android_binding_hint(row, status)
    return (
        a.replace("Content", "View")
        .replace("GroupQuickAddHub", "GroupQuickAddHubView")
        .replace("GroupActionRegistry", "GroupActionRegistry")
        .replace("AppShellScreen", "AppShellView")
    )


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    print("Loading field rows…")
    rows = load_field_rows()
    print(f"Field rows: {len(rows)}")
    print("Loading table index…")
    tables = load_table_index()
    print(f"Tables indexed: {len(tables)}")

    csv_headers = [
        "moment_family",
        "moment_subtype",
        "mapping_batch",
        "group_moment_type",
        "screen_id",
        "figma_screen",
        "figma_node",
        "widget_id",
        "figma_widget",
        "widget_type",
        "ui_label",
        "field",
        "table_column",
        "supporting_tables",
        "api_read",
        "api_write",
        "api_contract_excel",
        "excel_mapping_status",
        "android_binding",
        "ios_binding",
        "projection",
        "gx2_status",
        "closure_priority",
        "notes",
    ]

    out_rows = []
    status_counts = Counter()
    by_family_status = defaultdict(Counter)
    backlog = []

    for r in rows:
        status = classify_row(r)
        # If entity references unknown table and status would be WIRED with no live — leave as is
        entity = norm(r["Canonical Entity / Projection"])
        fields = norm(r["Primary Field(s)"])
        api = normalize_api(r["API / Service Contract"])
        direction = norm(r["Data Direction"]).upper()

        api_read = ""
        api_write = ""
        if api.startswith("GET") or "GET /" in api:
            api_read = api
        if any(api.startswith(p) for p in ("POST", "PATCH", "PUT", "DELETE")) or "POST /" in api or "PATCH /" in api:
            api_write = api
        if direction == "READ" and api and not api_read:
            api_read = api
        if direction in {"WRITE", "READ/WRITE"} and api and not api_write:
            api_write = api

        family = family_from_batch(norm(r["Mapping Batch"]), norm(r["Group Moment Type"]))
        subtype = subtype_from(norm(r["Group Moment Type"]))
        projection = ""
        if "projection." in entity.lower():
            projection = entity

        rec = {
            "moment_family": family,
            "moment_subtype": subtype,
            "mapping_batch": norm(r["Mapping Batch"]),
            "group_moment_type": norm(r["Group Moment Type"]),
            "screen_id": norm(r["Screen ID"]),
            "figma_screen": norm(r["Functional Screen Name"]),
            "figma_node": norm(r["Figma Node"]),
            "widget_id": norm(r["Widget ID"]),
            "figma_widget": norm(r["Widget Name"]),
            "widget_type": norm(r["Widget Type"]),
            "ui_label": norm(r["UI Label / Display"]),
            "field": fields,
            "table_column": f"{entity} :: {fields}" if entity else fields,
            "supporting_tables": norm(r["Supporting Table(s)"]),
            "api_read": api_read,
            "api_write": api_write,
            "api_contract_excel": api,
            "excel_mapping_status": norm(r["Mapping Status"]),
            "android_binding": android_binding_hint(r, status),
            "ios_binding": ios_binding_hint(r, status),
            "projection": projection or (entity if entity.startswith("projection.") else ""),
            "gx2_status": status,
            "closure_priority": norm(r["Closure Priority"]),
            "notes": norm(r["Gap / Developer Notes"]),
        }
        out_rows.append(rec)
        status_counts[status] += 1
        by_family_status[family][status] += 1
        if status in {"CLIENT_FIX", "API_GAP", "SCHEMA_GAP"}:
            backlog.append(rec)

    # Write CSV
    with OUT_CSV.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=csv_headers)
        w.writeheader()
        w.writerows(out_rows)

    assert "UNKNOWN" not in status_counts
    assert sum(status_counts.values()) == len(out_rows)

    # Summary MD
    lines = []
    lines.append("# GX2 Field Matrix")
    lines.append("")
    lines.append("**Authority:** Figma `575:7980` + `Momentra_Group_217_Screen_4210_Widget_FINAL_Field_Level_Master_Mapping.xlsx` + `Momentra_Master_Table_Repository.xlsx`.")
    lines.append("")
    lines.append(f"**Rows:** {len(out_rows)} widgets. **Statuses:** no UNKNOWN.")
    lines.append("")
    lines.append(f"**Full CSV:** [`GX2_FIELD_MATRIX.csv`](GX2_FIELD_MATRIX.csv)")
    lines.append("")
    lines.append("## Status rollup")
    lines.append("")
    lines.append("| Status | Count | Meaning |")
    lines.append("|---|---:|---|")
    meanings = {
        "WIRED": "Figma widget ↔ client ↔ live API ↔ table.column",
        "CLIENT_FIX": "Live API + SQL exist; client binding/nav/state incomplete",
        "API_GAP": "Table exists; live router missing read and/or write",
        "SCHEMA_GAP": "Widget needs column/table that does not exist",
        "FIGMA_GAP": "Designed; no product/schema contract (e.g. AI)",
        "LOCAL_ONLY": "Device/chrome only; must not pretend to persist",
        "DEFERRED": "Explicitly out of GX-2",
    }
    for st in ["WIRED", "CLIENT_FIX", "API_GAP", "SCHEMA_GAP", "FIGMA_GAP", "LOCAL_ONLY", "DEFERRED"]:
        lines.append(f"| `{st}` | {status_counts.get(st, 0)} | {meanings[st]} |")
    lines.append("")
    lines.append("## By family")
    lines.append("")
    lines.append("| Family | WIRED | CLIENT_FIX | API_GAP | SCHEMA_GAP | FIGMA_GAP | LOCAL_ONLY | DEFERRED | Total |")
    lines.append("|---|---:|---:|---:|---:|---:|---:|---:|---:|")
    for fam in ["Shared Experience", "Shared Purchase", "Shared Living", "Common Group"]:
        c = by_family_status[fam]
        total = sum(c.values())
        lines.append(
            f"| {fam} | {c['WIRED']} | {c['CLIENT_FIX']} | {c['API_GAP']} | {c['SCHEMA_GAP']} | {c['FIGMA_GAP']} | {c['LOCAL_ONLY']} | {c['DEFERRED']} | {total} |"
        )
    lines.append("")
    lines.append("## Live API overrides applied (Excel was stale)")
    lines.append("")
    lines.append("These routes are live on `backend/typescript/src/api/v1/router.ts` and reclassify many Excel PARTIAL/CONTRACT GAP rows to `WIRED` or `CLIENT_FIX`:")
    lines.append("")
    for x in sorted(LIVE_READ | LIVE_WRITE):
        lines.append(f"- `{x}`")
    lines.append("")
    lines.append("## GX2-B priority (CLIENT_FIX with live backend)")
    lines.append("")
    cf = [r for r in backlog if r["gx2_status"] == "CLIENT_FIX"]
    # Dedupe by screen + widget theme
    lines.append(f"CLIENT_FIX widgets: **{len(cf)}**. Focus closures:")
    lines.append("")
    lines.append("| Theme | Example screens | Action |")
    lines.append("|---|---|---|")
    lines.append("| Settlement tour | Settlement Flow / Group Finance | Enable Finance CTA → `POST …/settlements`; record payment method only |")
    lines.append("| Split strategies | Add Expense | Submit PERCENTAGE/EXACT/SHARES (server already computes) |")
    lines.append("| Budget | Setup + Update Budget | Bind to `PATCH …/budget` (already live) |")
    lines.append("| Invites / participants | Setup Add People | Already mostly wired; fix remaining PARTIAL bindings |")
    lines.append("")
    lines.append("## GX2-C priority (API_GAP tables)")
    lines.append("")
    ag = [r for r in backlog if r["gx2_status"] == "API_GAP"]
    entity_counts = Counter()
    for r in ag:
        ent = r["table_column"].split(" :: ")[0].strip().lower()
        if ent:
            entity_counts[ent] += 1
    lines.append(f"API_GAP widgets: **{len(ag)}**. Top entities:")
    lines.append("")
    lines.append("| Entity | Widget count | Promote route |")
    lines.append("|---|---:|---|")
    route_map = {
        "collaboration.planning_item": "POST/GET planning-items",
        "collaboration.booking": "POST/GET bookings",
        "collaboration.group_update": "POST/GET updates",
        "shared.poll": "POST/GET polls",
        "collaboration.poll": "POST/GET polls → shared.poll",
        "collaboration.purchase_item": "POST/GET purchase-items",
        "collaboration.resident": "POST/GET residents",
        "memory.memory": "POST memories + GET memory facet",
        "work.task": "DEFERRED / work domain — map only if Excel forces",
        "work.goal": "DEFERRED unless required",
        "finance.budget": "PATCH budget already live — recheck CLIENT_FIX",
        "projection.group_life": "Populate projection writers",
        "projection.group_memory": "Populate projection writers",
        "projection.group_pulse": "Extend payload (open tasks)",
    }
    for ent, n in entity_counts.most_common(25):
        hint = next((v for k, v in route_map.items() if k in ent), "See GX2-C")
        lines.append(f"| `{ent}` | {n} | {hint} |")
    lines.append("")
    lines.append("## Explicitly out (DEFERRED / FIGMA_GAP / LOCAL_ONLY)")
    lines.append("")
    lines.append(f"- DEFERRED: {status_counts['DEFERRED']} (UPI/GPay/bank rails, chat, FX, Goal/Community product, remove-participant)")
    lines.append(f"- FIGMA_GAP: {status_counts['FIGMA_GAP']} (AI insights, invented health scores)")
    lines.append(f"- LOCAL_ONLY: {status_counts['LOCAL_ONLY']} (chrome + setup prefs that must not fake-persist)")
    lines.append(f"- SCHEMA_GAP: {status_counts['SCHEMA_GAP']} (do not invent columns)")
    lines.append("")
    lines.append("## Column contract")
    lines.append("")
    lines.append("CSV columns match the GX-2 plan:")
    lines.append("")
    lines.append("```text")
    lines.append("moment_family → moment_subtype → figma_screen → figma_widget → field")
    lines.append("→ table_column → api_read → api_write → android_binding → ios_binding")
    lines.append("→ projection → gx2_status")
    lines.append("```")
    lines.append("")
    lines.append("See also [`GX2_A_GAP_BACKLOG.md`](GX2_A_GAP_BACKLOG.md) for CLIENT_FIX + API_GAP screen rollup.")
    lines.append("")
    OUT_MD.write_text("\n".join(lines), encoding="utf-8")

    # Backlog by screen
    bl = []
    bl.append("# GX2-A Gap Backlog (CLIENT_FIX + API_GAP)")
    bl.append("")
    bl.append("Derived from `GX2_FIELD_MATRIX.csv`. Use to drive GX2-B and GX2-C.")
    bl.append("")
    by_screen = defaultdict(lambda: Counter())
    for r in backlog:
        key = (r["moment_family"], r["screen_id"], r["figma_screen"], r["figma_node"])
        by_screen[key][r["gx2_status"]] += 1
    bl.append("| Family | Screen ID | Screen | Node | CLIENT_FIX | API_GAP |")
    bl.append("|---|---|---|---|---:|---:|")
    for key, c in sorted(by_screen.items(), key=lambda x: (-(x[1]["CLIENT_FIX"] + x[1]["API_GAP"]), x[0][1])):
        fam, sid, name, node = key
        bl.append(f"| {fam} | `{sid}` | {name} | `{node}` | {c['CLIENT_FIX']} | {c['API_GAP']} |")
    bl.append("")
    OUT_GAP.write_text("\n".join(bl), encoding="utf-8")

    print("Wrote", OUT_CSV)
    print("Wrote", OUT_MD)
    print("Wrote", OUT_GAP)
    print("Status counts:", dict(status_counts))


if __name__ == "__main__":
    main()
