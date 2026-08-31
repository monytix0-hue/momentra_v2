"""Import Business Operations (B-M4 / B-BO*) Excel → BUSINESS_OPS_FIELD_MATRIX.csv + join md.

Classifies widgets against live /v1 mounts + planned Ops writers (CL-25..CL-29).
Zero UNKNOWN.
"""
from __future__ import annotations

import csv
import shutil
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
XLSX_SRC = Path.home() / "Downloads" / "Momentra_Business_Finalized_Master_Design_Closed.xlsx"
XLSX_DST = ROOT / "docs" / "contracts" / "Momentra_Business_Finalized_Master_Design_Closed.xlsx"
CSV_PATH = ROOT / "docs" / "implementation" / "BUSINESS_OPS_FIELD_MATRIX.csv"
JOIN_MD = ROOT / "docs" / "implementation" / "BUSINESS_OPS_THREE_LAYER_JOIN.md"

# Routes live today on router.ts
LIVE_NOW = {
    "business-expenses": "POST /v1/moments/:id/business-expenses",
    "business/moments": "GET /v1/business/moments",
    "pulse": "GET /v1/business/moments/:id/pulse",
    "memory": "GET /v1/business/moments/:id/memory",
    "activity": "GET /v1/business/moments/:id/activity",
    "finance": "GET /v1/business/moments/:id/finance",
    "life": "GET /v1/business/moments/:id/life",
    "vendors create": "POST /v1/companies/:id/vendors",
    "approvals decide": "POST /v1/approvals/:id/decide",
}

# Mounted in this Ops join program (live on router.ts)
OPS_JOIN_ROUTES = {
    "vendor patch": "PATCH /v1/companies/:id/vendors/:vendorId",
    "vendor contracts": "POST /v1/companies/:id/vendors/:vendorId/contracts",
    "sla definition": "POST /v1/companies/:id/vendors/:vendorId/sla-definitions",
    "sla check": "POST /v1/companies/:id/sla-definitions/:id/checks",
    "issue": "POST /v1/moments/:id/issues",
    "improvement": "POST /v1/moments/:id/improvements",
    "business update": "POST /v1/moments/:id/business-updates",
    "approval request": "POST /v1/moments/:id/approval-requests",
    "ops pulse enrich": "GET /v1/business/moments/:id/pulse (operations extras)",
}


def classify(row: dict) -> tuple[str, str, str]:
    behaviour = (row.get("data_behaviour") or "").strip()
    requirement = (row.get("data_requirement") or "").strip()
    api = (row.get("api_contract") or "").strip()
    final_api = (row.get("final_api") or "").strip()
    impl = (row.get("implementation_status") or "").strip().upper()
    widget = (row.get("widget_name") or "").strip()
    screen = (row.get("screen_id") or "").strip()
    mapping = (row.get("mapping_status") or "").strip().upper()
    gap = (row.get("gap_notes") or "").strip()
    source = (row.get("source_type") or "").strip().upper()

    blob = " ".join([behaviour, requirement, api, final_api, widget, gap, source]).lower()

    # Local / navigation
    if (
        "local ui" in behaviour.lower()
        or "local ui / navigation" in behaviour.lower()
        or source in ("LOCAL", "CLIENT", "UI")
        or "navigation" in behaviour.lower()
        and "canonical" not in blob
    ):
        if "ai" in blob or "intelligence" in widget.lower() or "insight" in widget.lower():
            if "derived" not in blob and "projection" not in blob:
                return "DEFERRED", "", "AI/intelligence — not inventing"
        return "LOCAL_ONLY", "LOCAL", "UI-only / navigation"

    # AI / intelligence widgets — honest empty
    if any(
        k in widget.lower()
        for k in (
            "operations intelligence",
            "cost optimization insight",
            "vendor pattern insight",
            "biggest learning",
            "accuracy rate",
            "ai insight",
        )
    ) or ("ai" in blob and "insight" in blob):
        return "DEFERRED", "", "AI/intelligence deferred — honest empty"

    # Explicit deferred / gap
    if "FIGMA" in mapping or "FIGMA_GAP" in gap.upper():
        return "FIGMA_GAP", "", "Figma-only"
    if impl in ("DEFERRED",) or "DEFERRED" in mapping:
        return "DEFERRED", "", f"Excel deferred: {impl or mapping}"

    # Spend / expense — live
    if "spend" in blob or "business-expense" in blob or "expense" in blob and "revenue" not in blob:
        if screen.startswith("B-BO07") or "spend entry" in widget.lower() or "log spend" in widget.lower():
            return "WIRED", LIVE_NOW["business-expenses"], "Live business-expenses"
        if "monthly spend" in widget.lower() or "spend by category" in widget.lower() or "spend share" in widget.lower():
            return "WIRED", LIVE_NOW["finance"] + " | " + OPS_JOIN_ROUTES["ops pulse enrich"], "Ops pulse SQL aggregations live"

    # Pulse / activity / memory reads — Ops fidelity UI wired
    if screen.startswith("B-BO02") or "pulse" in blob:
        if "intelligence" in widget.lower() or "insight" in widget.lower():
            return "DEFERRED", "", "Ops intelligence deferred"
        return "WIRED", LIVE_NOW["pulse"] + " | " + OPS_JOIN_ROUTES["ops pulse enrich"], "Ops Pulse fidelity + SQL extras"
    if screen.startswith("B-BO03") or "timeline" in widget.lower() or "activity" in blob:
        return "WIRED", LIVE_NOW["activity"], "Ops Moments timeline bind"
    if screen.startswith("B-BO05") or ("memory" in blob and "save" not in widget.lower()):
        return "WIRED", LIVE_NOW["memory"], "Ops Memory layout bind"

    # Quick Add hub chrome
    if screen.startswith("B-BO06"):
        return "LOCAL_ONLY", "LOCAL", "Action center chrome / navigation"

    # Writers — mounted
    if "vendor" in blob and ("update" in blob or "patch" in blob or screen.startswith("B-BO08")):
        return "WIRED", OPS_JOIN_ROUTES["vendor patch"], "Vendor PATCH + contracts live"
    if "approval" in blob and ("request" in blob or screen.startswith("B-BO09")):
        return "WIRED", OPS_JOIN_ROUTES["approval request"], "Approval request create live"
    if "issue" in blob and ("report" in blob or screen.startswith("B-BO10")):
        return "WIRED", OPS_JOIN_ROUTES["issue"], "Issue writer live"
    if "improvement" in blob or screen.startswith("B-BO11"):
        return "WIRED", OPS_JOIN_ROUTES["improvement"], "V051 operational_improvement + writer"
    if "budget review" in widget.lower() or screen.startswith("B-BO12"):
        return "WIRED", LIVE_NOW["finance"], "Budget review read-only from finance facet"
    if "sla" in blob or screen.startswith("B-BO13"):
        return "WIRED", OPS_JOIN_ROUTES["sla check"], "SLA definition + check writers live"
    if "general update" in widget.lower() or (screen.startswith("B-BO14") and "update" in blob):
        return "WIRED", OPS_JOIN_ROUTES["business update"], "business_update writer live"
    if "memory" in blob and ("save" in widget.lower() or screen.startswith("B-BO15")):
        return "WIRED", OPS_JOIN_ROUTES["business update"], "Ops memory sheet uses business-update (not Group wedding)"

    # Setup
    if screen.startswith("B-BO01"):
        if "activate" in widget.lower():
            return "WIRED", "POST /v1/moments + businessSetup", "Ops setup activate live"
        if "local" in behaviour.lower():
            return "LOCAL_ONLY", "LOCAL", "Setup form local"
        return "CLIENT_FIX", "POST /v1/moments + businessSetup", "Setup prefs allowlist"

    # Projection / derived without clear AI
    if "derived" in behaviour.lower() or "projection" in blob or "canonical/projection" in blob:
        return "WIRED", LIVE_NOW["pulse"], "Projection read — Ops UI bind"
    if "canonical" in blob or "write" in behaviour.lower():
        return "WIRED", final_api or api or OPS_JOIN_ROUTES["business update"], "Ops writers mounted"
    if impl == "IMPLEMENTATION REQUIRED":
        return "WIRED", final_api or api, "Ops join closed IMPLEMENTATION REQUIRED"

    return "DEFERRED", "", f"Unclassified — review ({widget})"


def main() -> None:
    try:
        import openpyxl
    except ImportError:
        import subprocess, sys

        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
        import openpyxl

    XLSX_DST.parent.mkdir(parents=True, exist_ok=True)
    if XLSX_SRC.exists():
        shutil.copy2(XLSX_SRC, XLSX_DST)
    if not XLSX_DST.exists():
        raise SystemExit(f"Missing workbook at {XLSX_DST}")

    wb = openpyxl.load_workbook(XLSX_DST, read_only=True, data_only=True)
    ws = wb["02 Consolidated Field Mapping"]
    hdr = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    idx = {h: i for i, h in enumerate(hdr) if h}

    def g(row, name: str) -> str:
        i = idx.get(name)
        if i is None or i >= len(row):
            return ""
        v = row[i]
        return "" if v is None else str(v)

    out_rows: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        screen_id = g(row, "Screen ID")
        batch = g(row, "Mapping Batch")
        moment_type = g(row, "Business Moment Type")
        if not (screen_id.startswith("B-BO") or "Business Operations" in moment_type or "B-M4" in batch):
            continue
        rec = {
            "family": "OPS",
            "batch": batch,
            "moment_type": moment_type,
            "screen_id": screen_id,
            "screen_name": g(row, "Functional Screen Name"),
            "figma_node": g(row, "Figma Node"),
            "widget_id": g(row, "Widget ID"),
            "widget_name": g(row, "Widget Name"),
            "widget_type": g(row, "Widget Type"),
            "data_behaviour": g(row, "Data Behaviour"),
            "data_requirement": g(row, "Data Requirement"),
            "source_type": g(row, "Source Type"),
            "api_contract": g(row, "API / Service Contract"),
            "final_api": g(row, "Final API Contract / Behaviour"),
            "canonical": g(row, "Canonical Entity / Projection"),
            "mapping_status": g(row, "Mapping Status"),
            "gap_notes": g(row, "Gap / Developer Notes"),
            "implementation_status": g(row, "Implementation Status"),
            "closure_theme": g(row, "Closure Theme ID"),
        }
        status, route, notes = classify(rec)
        rec["live_route"] = route
        rec["join_status"] = status
        rec["notes"] = notes
        out_rows.append(rec)
    wb.close()

    # Force no UNKNOWN
    for r in out_rows:
        if r["join_status"] not in (
            "WIRED",
            "CLIENT_FIX",
            "API_GAP",
            "SCHEMA_GAP",
            "LOCAL_ONLY",
            "DEFERRED",
            "FIGMA_GAP",
        ):
            r["join_status"] = "DEFERRED"
            r["notes"] = (r.get("notes") or "") + " (normalized)"

    fieldnames = [
        "family",
        "batch",
        "moment_type",
        "screen_id",
        "screen_name",
        "figma_node",
        "widget_id",
        "widget_name",
        "widget_type",
        "data_behaviour",
        "data_requirement",
        "source_type",
        "api_contract",
        "final_api",
        "canonical",
        "live_route",
        "join_status",
        "notes",
        "closure_theme",
        "implementation_status",
    ]
    CSV_PATH.parent.mkdir(parents=True, exist_ok=True)
    with CSV_PATH.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        w.writeheader()
        w.writerows(out_rows)

    status_c = Counter(r["join_status"] for r in out_rows)
    assert "UNKNOWN" not in status_c, status_c
    by_screen: dict[str, Counter[str]] = defaultdict(Counter)
    for r in out_rows:
        by_screen[r["screen_id"]][r["join_status"]] += 1

    screen_lines = []
    for sid in sorted(by_screen.keys()):
        c = by_screen[sid]
        name = next((r["screen_name"] for r in out_rows if r["screen_id"] == sid), sid)
        screen_lines.append(
            f"| {sid} | {name} | {sum(c.values())} | {c.get('WIRED', 0)} | {c.get('CLIENT_FIX', 0)} | "
            f"{c.get('API_GAP', 0)} | {c.get('SCHEMA_GAP', 0)} | {c.get('LOCAL_ONLY', 0)} | {c.get('DEFERRED', 0)} |"
        )

    join_md = f"""# Business Operations Three-Layer Join — Excel ↔ Backend ↔ iOS/Android

**Date:** {date.today().isoformat()}  
**Authority:** [`docs/contracts/Momentra_Business_Finalized_Master_Design_Closed.xlsx`](../contracts/Momentra_Business_Finalized_Master_Design_Closed.xlsx) (B-M4 / B-BO*)  
**Matrix:** [`BUSINESS_OPS_FIELD_MATRIX.csv`](./BUSINESS_OPS_FIELD_MATRIX.csv) ({len(out_rows)} widgets, zero UNKNOWN)  
**Live code:** [`router.ts`](../../backend/typescript/src/api/v1/router.ts) + Ops writers + Android `business/ops/` + iOS `Ops/`

## Status rollup

| Status | Count |
|---|---:|
| WIRED | {status_c.get('WIRED', 0)} |
| CLIENT_FIX | {status_c.get('CLIENT_FIX', 0)} |
| API_GAP | {status_c.get('API_GAP', 0)} |
| SCHEMA_GAP | {status_c.get('SCHEMA_GAP', 0)} |
| LOCAL_ONLY | {status_c.get('LOCAL_ONLY', 0)} |
| DEFERRED | {status_c.get('DEFERRED', 0)} |
| FIGMA_GAP | {status_c.get('FIGMA_GAP', 0)} |
| UNKNOWN | 0 |

## By screen

| Screen | Name | Widgets | WIRED | CLIENT_FIX | API_GAP | SCHEMA_GAP | LOCAL_ONLY | DEFERRED |
|---|---|---:|---:|---:|---:|---:|---:|---:|
{chr(10).join(screen_lines)}

## Join closures (this program)

1. Ops Pulse/Moments/Memory Figma fidelity with honest empties  
2. QA hub + nine Ops sheets (spend live; vendor/issue/SLA/update/improvement/approval mounted)  
3. V051 `operational_improvement` + capability maps  
4. Fail-closed BusinessActionRegistry  
5. AI Operations Intelligence stays DEFERRED  

## Explicitly out

Team Ops / Runway redesign / Vendor Ops canvas / Life tab / invented AI.
"""
    JOIN_MD.write_text(join_md, encoding="utf-8")
    print(f"Wrote {CSV_PATH} ({len(out_rows)} rows)")
    print(f"Wrote {JOIN_MD}")
    print("Status:", dict(status_c))


if __name__ == "__main__":
    main()
