from openpyxl import load_workbook
from pathlib import Path

files = [
    Path(r"c:\Users\HI\Downloads\Momentra_Group_217_Screen_4210_Widget_FINAL_Field_Level_Master_Mapping.xlsx"),
    Path(r"c:\Users\HI\Downloads\Momentra_Master_Table_Repository.xlsx"),
]
for f in files:
    print("=" * 80)
    print(f.name, "exists=", f.exists(), "size=", f.stat().st_size if f.exists() else None)
    if not f.exists():
        continue
    wb = load_workbook(f, read_only=True, data_only=True)
    print("sheets:", wb.sheetnames)
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        print(f"\n--- Sheet: {name} | rows={len(rows)} ---")
        for i, row in enumerate(rows[:12]):
            print(i, row)
    wb.close()
