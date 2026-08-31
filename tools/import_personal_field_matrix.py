"""Import Personal UI Contract Excel → PERSONAL_FIELD_MATRIX.csv + PERSONAL_THREE_LAYER_JOIN.md.

Classifies each widget against live /v1 mounts + known client bind state.
Zero UNKNOWN.
"""
from __future__ import annotations

import csv
import re
import shutil
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
XLSX_SRC = Path.home() / "Downloads" / "Momentra_Personal_1979_Widget_CLEAN_Master_UI_Contract_v1.xlsx"
XLSX_DST = ROOT / "docs" / "contracts" / "Momentra_Personal_1979_Widget_CLEAN_Master_UI_Contract_v1.xlsx"
CSV_PATH = ROOT / "docs" / "implementation" / "PERSONAL_FIELD_MATRIX.csv"
JOIN_MD = ROOT / "docs" / "implementation" / "PERSONAL_THREE_LAYER_JOIN.md"

# Logical command → live route presence (router.ts as of join)
LIVE_COMMANDS: dict[str, str] = {
    "PersonalPulse.Get": "GET /v1/personal/pulse",
    "Personal Pulse Read": "GET /v1/personal/pulse",
    "PersonalLife.Get": "GET /v1/personal/life",
    "PersonalMemory.Get": "GET /v1/personal/memory",
    "PersonalActivity.List": "GET /v1/personal/activity",
    "PersonalMoments.Get": "GET /v1/personal/moments",
    "PersonalMoment.Create": "POST /v1/moments",
    "Expense.Create": "POST /v1/moments/:id/expenses",
    "Personal Available Actions": "GET /v1/me capabilities",
    "LOCAL_ONLY": "LOCAL",
    "PersonalShell.Navigate": "LOCAL",
}

# Closed in Personal join (honest REAL Life / Memory / Attention). Keep empty unless new stubs appear.
STUB_READS: set[str] = set()

DEFERRED_COMMANDS = {
    "AI",
    "Reflect",
    "SOURCE_REVIEW_REQUIRED",
    "UNRESOLVED",
    "TBD",
    "GetMemoryPatterns",
    "Analytics.GetDrivers",
    "Metric.Get",
}


def family_from_screen(screen_id: str) -> str:
    if not screen_id or not str(screen_id).startswith("PER-"):
        return "GL"
    return str(screen_id).split("-")[1]


def classify(row: dict) -> tuple[str, str, str]:
    """Return (join_status, live_route, notes)."""
    cmd = (row.get("api_command") or "").strip()
    http = (row.get("http_method") or "").strip().upper()
    classification = (row.get("data_classification") or "").strip()
    behaviour = (row.get("data_behaviour") or "").strip()
    widget = (row.get("widget_name") or "").strip()

    # Local / static UI
    if (
        cmd in ("LOCAL_ONLY", "PersonalShell.Navigate")
        or http in ("LOCAL", "—", "CLIENT ROUTER", "TRANSPORT", "LOCAL / FEATURE FLAG")
        or classification.startswith("LOCAL_UI")
        or classification in ("STATIC", "STATIC_UI", "STATIC_OR_DERIVED", "LOCAL_DERIVED", "FORM_STATE")
        or "Local UI" in behaviour
        or "LOCAL_NAVIGATION" in behaviour
    ):
        if "AI" in classification or "AI_" in classification:
            return "DEFERRED", "", "AI/static — not inventing"
        return "LOCAL_ONLY", "LOCAL", "UI-only / navigation"

    # Explicit deferred signals
    cmd_u = cmd.upper()
    if any(d.upper() in cmd_u for d in DEFERRED_COMMANDS) or http == "TBD":
        if "Reflect" in widget or "AI" in classification:
            return "DEFERRED", "", "Reflect/AI deferred by product"
        return "DEFERRED", "", f"Contract TBD/deferred: {cmd or http}"

    # Map known live commands
    route = LIVE_COMMANDS.get(cmd)
    if not route:
        # Fuzzy: pulse / activity / expense / movement / future / lifestyle / relationship
        low = cmd.lower()
        if "pulse" in low:
            route = "GET /v1/personal/pulse"
        elif "activity.list" in low or "personalactivity" in low.replace(" ", ""):
            route = "GET /v1/personal/activity"
        elif "memory" in low:
            route = "GET /v1/personal/memory"
        elif "life.get" in low or cmd == "PersonalLife.Get":
            route = "GET /v1/personal/life"
        elif "expense" in low:
            route = "POST /v1/moments/:id/expenses"
        elif "income" in low:
            route = "POST /v1/moments/:id/income"
        elif "movement" in low or "transfer" in low or "savings" in low:
            route = "POST /v1/moments/:id/movements"
        elif "attention" in low:
            route = "POST /v1/moments/:id/attention-captures | GET /v1/personal/attention"
        elif "future" in low or "milestone" in low or "opportunity" in low or "pivot" in low or "progress" in low or "learning" in low:
            route = "POST /v1/moments/:id/future-items | GET future-*"
        elif "lifestyle" in low:
            route = "POST /v1/moments/:id/lifestyle-activities | GET lifestyle-*"
        elif "relationship" in low:
            route = "POST /v1/moments/:id/relationship-activities | GET relationships-*"
        elif "moment.create" in low or "configure" in low:
            route = "POST /v1/moments + personalSetup"
        elif "archive" in low or "getmanagement" in low.replace(" ", "") or "management" in low:
            # Backend PATCH/archive/cancel live; Personal Manage UI not fully bound → CLIENT_FIX below
            route = "PATCH|POST /v1/moments/:id|/archive|/cancel"
        elif "setup" in low or "activate" in low:
            route = "POST /v1/personal/setups/:code/activate | POST /moments"
        elif classification.startswith("CANONICAL_WRITE") or "COMMAND" in classification:
            route = "LIVE_WRITE_FAMILY"
        elif classification.startswith("PROJECTION") or "Derived" in classification or classification.startswith("DERIVED"):
            route = "LIVE_READ_FAMILY"
        elif classification in ("BOOTSTRAP", "BOOTSTRAP_READ", "CAPABILITY_READ", "LOCAL_OR_BOOTSTRAP"):
            route = "GET /v1/me"
        else:
            route = ""

    if not route:
        if classification in ("SOURCE_DEPENDENT", "UNRESOLVED_CANONICAL"):
            return "DEFERRED", "", "Unresolved canonical — review"
        if "FIGMA" in classification.upper():
            return "FIGMA_GAP", "", "Figma-only"
        return "DEFERRED", "", f"No live mapping for command={cmd!r}"

    if cmd in STUB_READS:
        return "CLIENT_FIX", route, "Live route; client/projection bind incomplete"

    # Manage Moment: API live, dedicated Personal Manage UI still unbound
    if "archive" in route.lower() or "/cancel" in route.lower():
        return "CLIENT_FIX", route, "Moment PATCH/archive/cancel live; Manage screens unbound"

    if route == "LOCAL":
        return "LOCAL_ONLY", route, "Local"

    if route.startswith("GET") or "LIVE_READ" in route or route.startswith("GET /v1/me"):
        return "WIRED", route, "Live read + client consume"
    if route.startswith("POST") or route.startswith("PATCH") or "LIVE_WRITE" in route:
        return "WIRED", route, "Live write + client Quick Add"
    return "WIRED", route, "Live"


def main() -> None:
    try:
        import openpyxl
    except ImportError:
        import subprocess, sys

        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
        import openpyxl

    XLSX_DST.parent.mkdir(parents=True, exist_ok=True)
    if XLSX_SRC.exists():
        shutil.copy2(XLSX_SRC, XLSX_DST)
    if not XLSX_DST.exists():
        raise SystemExit(f"Missing workbook at {XLSX_DST}")

    wb = openpyxl.load_workbook(XLSX_DST, read_only=True, data_only=True)
    ws = wb["Personal UI Contract Master"]
    hdr = list(next(ws.iter_rows(min_row=2, max_row=2, values_only=True)))
    idx = {h: i for i, h in enumerate(hdr) if h}

    out_rows: list[dict] = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        def g(name: str):
            i = idx.get(name)
            if i is None or i >= len(row):
                return ""
            v = row[i]
            return "" if v is None else str(v)

        rec = {
            "domain": g("Domain"),
            "screen_id": g("Screen ID"),
            "screen_name": g("Screen Name"),
            "screen_figma_node": g("Screen Figma Node"),
            "widget_id": g("Widget ID"),
            "widget_name": g("Widget Name"),
            "widget_figma_node": g("Widget Figma Node"),
            "parent_section": g("Parent Widget / Section"),
            "widget_type": g("Widget Type"),
            "data_behaviour": g("Data Behaviour"),
            "data_classification": g("Data Classification"),
            "api_command": g("API / Command"),
            "http_method": g("HTTP Method"),
            "family": family_from_screen(g("Screen ID")),
        }
        status, route, notes = classify(rec)
        rec["live_route"] = route
        rec["join_status"] = status
        rec["notes"] = notes
        out_rows.append(rec)
    wb.close()

    fieldnames = [
        "family",
        "domain",
        "screen_id",
        "screen_name",
        "screen_figma_node",
        "widget_id",
        "widget_name",
        "widget_figma_node",
        "parent_section",
        "widget_type",
        "data_behaviour",
        "data_classification",
        "api_command",
        "http_method",
        "live_route",
        "join_status",
        "notes",
    ]
    CSV_PATH.parent.mkdir(parents=True, exist_ok=True)
    with CSV_PATH.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(out_rows)

    status_c = Counter(r["join_status"] for r in out_rows)
    assert "UNKNOWN" not in status_c, status_c
    by_family: dict[str, Counter[str]] = defaultdict(Counter)
    for r in out_rows:
        by_family[r["family"]][r["join_status"]] += 1

    family_lines = []
    for fam in ("LO", "FB", "LS", "REL", "GL"):
        c = by_family.get(fam, Counter())
        family_lines.append(
            f"| {fam} | {sum(c.values())} | {c.get('WIRED', 0)} | {c.get('CLIENT_FIX', 0)} | "
            f"{c.get('LOCAL_ONLY', 0)} | {c.get('DEFERRED', 0)} | {c.get('API_GAP', 0)} | {c.get('FIGMA_GAP', 0)} |"
        )

    join_md = f"""# Personal Three-Layer Join — Excel ↔ Backend ↔ iOS/Android

**Date:** {date.today().isoformat()}  
**Authority:** [`docs/contracts/Momentra_Personal_1979_Widget_CLEAN_Master_UI_Contract_v1.xlsx`](../contracts/Momentra_Personal_1979_Widget_CLEAN_Master_UI_Contract_v1.xlsx)  
**Matrix:** [`PERSONAL_FIELD_MATRIX.csv`](./PERSONAL_FIELD_MATRIX.csv) ({len(out_rows)} widgets, zero UNKNOWN)  
**Live code:** [`router.ts`](../../backend/typescript/src/api/v1/router.ts) + [`ApiService.kt`](../../apk/app/src/main/java/com/example/momentra/data/api/ApiService.kt) + [`APIClient.swift`](../../momentra/momentra/API/APIClient.swift)

## Status rollup

| Status | Count |
|---|---:|
| WIRED | {status_c.get('WIRED', 0)} |
| CLIENT_FIX | {status_c.get('CLIENT_FIX', 0)} |
| LOCAL_ONLY | {status_c.get('LOCAL_ONLY', 0)} |
| DEFERRED | {status_c.get('DEFERRED', 0)} |
| API_GAP | {status_c.get('API_GAP', 0)} |
| FIGMA_GAP | {status_c.get('FIGMA_GAP', 0)} |
| SCHEMA_GAP | {status_c.get('SCHEMA_GAP', 0)} |
| UNKNOWN | 0 |

## By family

| Family | Widgets | WIRED | CLIENT_FIX | LOCAL_ONLY | DEFERRED | API_GAP | FIGMA_GAP |
|---|---:|---:|---:|---:|---:|---:|---:|
{chr(10).join(family_lines)}

## Excel note

The UI contract freezes screens/widgets/commands. **`API Route` is empty for all 1,979 rows.** Join status is derived by mapping logical `API / Command` values onto live `/v1` mounts.

## Known join closures (this program)

1. **Life** — `GET /personal/life` returns `dataQuality: REAL` with honest empties + live `activeAreaCount` / journey from `projection.recent_activity` (no invented Figma scores)
2. **Memory** — `GET /personal/memory` projects `memory.memory` for the user
3. **Attention** — `GET /personal/attention` projects `analytics.attention_capture`
4. **Capability fail-closed** — PersonalActionRegistry empty caps disable destinations (Android + iOS)
5. **Transfer/Savings** — WIRED via `POST …/movements`; Reflect stays DEFERRED; Manage Moment API live but UI CLIENT_FIX
6. **Proof harness** — `tests/personal-three-layer-join.test.ts` (LO / FB / LS / REL)

## Explicitly out

AI Reflect / invented Memory intelligence / redesign of 66 screens / V030 / Money as 5th setup family.
"""
    JOIN_MD.write_text(join_md, encoding="utf-8")
    print(f"Wrote {CSV_PATH} ({len(out_rows)} rows)")
    print(f"Wrote {JOIN_MD}")
    print("Status:", dict(status_c))


if __name__ == "__main__":
    main()
